from filters import AdminsMessage, IsSubscriber
from handlers.users.levels import set_levels
from keyboards.cancel import keyboard_cancel
from loader import dp
from aiogram import types
from aiogram.dispatcher import FSMContext
from aiogram.dispatcher.filters import Command
import os
from misc.save_excel import create_excel_file
from utils.db_api import users_commands as commands
from asyncio import sleep
from aiogram.dispatcher.filters.state import StatesGroup, State
from loguru import logger
from openpyxl import Workbook
from aiogram import Dispatcher
from data.config import admins
from aiogram.types import InputFile

from utils.db_api.admin_commands import get_newsletter_text, replace_newsletter_period, get_set_bonus_1, \
    get_set_bonus_2, set_set_bonus_1, set_set_bonus_2, get_tariff, set_tariff, get_variants_proposal, \
    set_variants_proposal
from utils.db_api.proposal_commands import update_status_by_id, get_proposals_by_status, get_all_proposals
from utils.db_api.users_commands import change_user_role, select_all_users_with_data, get_user_referrals, \
    find_user_by_referral_value, count_users, save_count_levels, get_user_id_by_username, get_user_created_at
from aiogram.types import ReplyKeyboardMarkup, KeyboardButton, ReplyKeyboardRemove


class SendMessage(StatesGroup):
    send_mess = State()
    time = State()
    mess = State()
    user_id = State()
    role = State()
    bonus_selection = State()
    amount = State()
    tariff = State()
    options = State()
    applications_selection = State()
    number_bid = State()
    status = State()
    username = State()
    username_mess = State()
    service_agreement = State()


@dp.message_handler(IsSubscriber(), text="/test")
async def command_help(message: types.Message):
    await message.answer(f'Тест ')
    created_at = await get_user_created_at(1089138631)
    print(created_at)


@dp.message_handler(AdminsMessage(), Command('admin'))
async def req(message: types.Message):
    await message.answer(f'/message - отправить сообщения всем пользователям\n'
                         f'/message_user - отправить сообщение пользователю по id или username \n'
                         f'/auto_mailing - создать рассылку сообщений по времени\n'
                         f'/set_user_status - задать статус пользователю (роль)\n'
                         f'/edit_bonuses - редактировать процент бонусов\n'
                         f'/edit_tariff - редактировать тариф \n'
                         f'/set_status_bid - установить статус заявки\n'
                         f'/edit_bid_options - редактировать варианты заявок \n'
                         f'/list_users - выгрузить пользователей из БД\n'
                         f'/change_user_list - загрузить изменения пользователей в БД\n'
                         f'/unload_bid - выгрузить заявки \n'
                         f'/update_referrals - пересчитать рефералы пользователям\n'
                         f'/ref_id - узнать сколько у юзера пользователей по id \n'
                         f'/delete_user - удалить пользователя по id \n'
                         f'/replace_service_agreement - заменить договор оказания услуг'
                         )


# команда /message отправить сообщения всем пользователям
@dp.message_handler(AdminsMessage(), Command('message'))
async def send_message(message: types.Message):
    await message.answer('Какое сообщение отправить:', reply_markup=keyboard_cancel)
    await SendMessage.send_mess.set()  # устанавливаем состояние


# Отлавливаем состояние введенного сообщения
@dp.message_handler(state=SendMessage.send_mess)
async def send_message(message: types.Message, state: FSMContext):
    text = message.text
    if text.lower() == 'отмена':
        await message.answer('Отменено', reply_markup=ReplyKeyboardRemove())
        await state.finish()  # обязательно завершаем состояние
    else:
        await send_mess(text)
        await message.answer('отправлено')
        await state.finish()


# команда /message отправить сообщения конкретному пользователю
@dp.message_handler(AdminsMessage(), Command('message_user'))
async def send_message(message: types.Message):
    await message.answer('Введите id или username пользователя:', reply_markup=keyboard_cancel)
    await SendMessage.username.set()  # устанавливаем состояние


# Отлавливаем состояние введенного сообщения
@dp.message_handler(state=SendMessage.username)
async def username(message: types.Message, state: FSMContext):
    text = message.text
    if text.lower() == 'отмена':
        await message.answer('Отменено', reply_markup=ReplyKeyboardRemove())
        await state.finish()  # обязательно завершаем состояние
    else:
        user_id = None

        try:
            # Попытка преобразовать текст в число
            user_id = int(text)
        except ValueError:
            # Если не удалось преобразовать в число, проверяем наличие символа '@'
            if text.startswith('@'):
                user_id = await get_user_id_by_username(text[1:])

        if user_id is not None:
            # Общие действия после успешной проверки идентификатора
            await state.update_data(user_id=user_id)
            await message.answer(f'Введите сообщение для {text}', reply_markup=keyboard_cancel)
            await SendMessage.username_mess.set()
        else:
            await message.answer(f'Введите либо id "987654321", либо username "@Username"',
                                 reply_markup=keyboard_cancel)


# Отлавливаем состояние введенного id
@dp.message_handler(state=SendMessage.username_mess)
async def send_message_id(message: types.Message, state: FSMContext):
    mess = message.text
    if mess.lower() == 'отмена':
        await message.answer('Отменено', reply_markup=ReplyKeyboardRemove())
        await state.finish()  # обязательно завершаем состояние
    else:
        data = await state.get_data()
        user_id = data['user_id']
        await dp.bot.send_message(user_id, f'Сообщение от админа: <b>{mess}</b>')
        await message.answer(f'Сообщение отправлено!')
        await state.finish()


# команда /auto_mailing отправить сообщения всем пользователям по времени
@dp.message_handler(AdminsMessage(), Command('auto_mailing'))
async def send_auto_mailing(message: types.Message):
    await message.answer('Установите частоту отправки (введите количество дней, например "3" - '
                         'отправка будет 1 раз в 3 дня):', reply_markup=keyboard_cancel)
    await SendMessage.time.set()  # устанавливаем состояние


# Отлавливаем состояние введенного day
@dp.message_handler(state=SendMessage.time)
async def input_day(message: types.Message, state: FSMContext):
    text = message.text
    if text.lower() == 'отмена':
        await message.answer('Отменено', reply_markup=ReplyKeyboardRemove())
        await state.finish()  # обязательно завершаем состояние
    else:
        await state.update_data(day=message.text)
        await message.answer('Какое сообщение отправлять:', reply_markup=keyboard_cancel)
        await SendMessage.mess.set()  # устанавливаем состояние


# Отлавливаем состояние введенного сообщения
@dp.message_handler(state=SendMessage.mess)
async def send_mailing(message: types.Message, state: FSMContext):
    text = message.text
    if text.lower() == 'отмена':
        await message.answer('Отменено', reply_markup=ReplyKeyboardRemove())
        await state.finish()  # обязательно завершаем состояние
    else:
        data = await state.get_data()
        day = data['day']
        await replace_newsletter_period(int(day))
        await message.answer(f'Каждые {day} дня, отправлять сообщение {message.text}')
        await state.finish()


async def send_mess(message: types.Message):
    # все пользователи из БД
    users = await commands.select_all_users()
    for user in users:
        try:
            await dp.bot.send_message(chat_id=user.user_id, text=message)
            await sleep(0.25)  # 4 сообщения в секунду
        except Exception as e:
            logger.error(e)


# команда /set_user_status задать статус пользователю (роль)
@dp.message_handler(AdminsMessage(), Command('set_user_status'))
async def set_user_status(message: types.Message):
    await message.answer('Введите user_id пользователя:', reply_markup=keyboard_cancel)
    await SendMessage.user_id.set()  # устанавливаем состояние


# Отлавливаем состояние введенного user_id
@dp.message_handler(state=SendMessage.user_id)
async def input_data(message: types.Message, state: FSMContext):
    text = message.text
    if text.lower() == 'отмена':
        await message.answer('Отменено', reply_markup=ReplyKeyboardRemove())
        await state.finish()  # обязательно завершаем состояние
    else:
        await state.update_data(user_id=message.text)

        # Создаем клавиатуру с кнопками ролей
        keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True)
        roles = ['admin', 'handler', 'manager', 'copywriter', 'client', 'отмена']
        keyboard.add(*roles)

        await message.answer('Выберите роль пользователя:', reply_markup=keyboard)
        await SendMessage.role.set()  # устанавливаем состояние


@dp.message_handler(state=SendMessage.role)
async def input_role(message: types.Message, state: FSMContext):
    text = message.text
    if text.lower() == 'отмена':
        await message.answer('Отменено', reply_markup=ReplyKeyboardRemove())
        await state.finish()  # обязательно завершаем состояние
    else:
        role = message.text
        data = await state.get_data()
        user_id = int(data['user_id'])
        # функция записи данных в БД
        await state.finish()
        # убрать кнопки
        keyboard = types.ReplyKeyboardRemove()

        if await change_user_role(user_id=user_id, new_role=role):
            await message.answer(f'Пользователю {user_id} установлен статус: {role}', reply_markup=keyboard)
        else:
            await message.answer(f'Пользователя с {user_id} нет!', reply_markup=keyboard)


# edit_bonuses - редактировать процент бонусов
@dp.message_handler(AdminsMessage(), Command('edit_bonuses'))
async def input_selection(message: types.Message):
    # Создаем клавиатуру с кнопками ролей
    keyboard_bonuses = types.ReplyKeyboardMarkup(resize_keyboard=True)
    roles = ['ожидаемый', 'фактический', 'отмена']
    keyboard_bonuses.add(*roles)
    await message.answer('Выберите бонусную программу:', reply_markup=keyboard_bonuses)
    await SendMessage.bonus_selection.set()  # устанавливаем состояние


# Отлавливаем состояние бонусную программ
@dp.message_handler(state=SendMessage.bonus_selection)
async def input_bonuses(message: types.Message, state: FSMContext):
    text = message.text
    if text.lower() == 'отмена':
        await message.answer('Отменено', reply_markup=ReplyKeyboardRemove())
        await state.finish()  # обязательно завершаем состояние
    else:
        bonus_selection = message.text
        await state.update_data(bonus_selection=bonus_selection)
        get_bonus = ""
        if bonus_selection == "ожидаемый":
            get_bonus = await get_set_bonus_1()
        elif bonus_selection == "фактический":
            get_bonus = await get_set_bonus_2()
        else:
            await message.answer(f'Введено неверное значение"', reply_markup=keyboard_cancel)
        await message.answer(f'{bonus_selection} бонус равен {get_bonus}\n'
                             f'Введите новое значение:', reply_markup=keyboard_cancel)
        await SendMessage.amount.set()  # устанавливаем состояние


# Отлавливаем состояние введенного сообщения
@dp.message_handler(state=SendMessage.amount)
async def input_amount(message: types.Message, state: FSMContext):
    text = message.text
    if text.lower() == 'отмена':
        await message.answer('Отменено', reply_markup=ReplyKeyboardRemove())
        await state.finish()  # обязательно завершаем состояние
    else:
        try:
            amount = int(message.text)
        except ValueError:
            await message.answer("Пожалуйста, введите число:")
            return
        data = await state.get_data()
        bonus_selection = data['bonus_selection']
        if bonus_selection == "ожидаемый":
            await set_set_bonus_1(amount)
        elif bonus_selection == "фактический":
            await set_set_bonus_2(amount)
        await message.answer("Установлено!")
        await state.finish()


# /edit_rate - редактировать тариф
@dp.message_handler(AdminsMessage(), Command('edit_tariff'))
async def edit_rate(message: types.Message):
    tariff = await get_tariff()
    await message.answer(f'Тариф равен {tariff}, установите новое значение:', reply_markup=keyboard_cancel)
    await SendMessage.tariff.set()  # устанавливаем состояние


# Отлавливаем состояние введенного сообщения
@dp.message_handler(state=SendMessage.tariff)
async def input_tariff(message: types.Message, state: FSMContext):
    text = message.text
    if text.lower() == 'отмена':
        await message.answer('Отменено', reply_markup=ReplyKeyboardRemove())
        await state.finish()  # обязательно завершаем состояние
    else:
        try:
            tariff = int(message.text)
        except ValueError:
            await message.answer("Пожалуйста, введите число:")
            return
        await set_tariff(tariff)
        await message.answer('Установлено!')
        await state.finish()


# f'/set_status_bid - установить статус заявки
@dp.message_handler(AdminsMessage(), Command('set_status_bid'))
async def input_set_status_bid(message: types.Message):
    await message.answer('Введите номер заявки:', reply_markup=keyboard_cancel)
    await SendMessage.number_bid.set()  # устанавливаем состояние


# Отлавливаем состояние введенного number_bid
@dp.message_handler(state=SendMessage.number_bid)
async def input_data(message: types.Message, state: FSMContext):
    text = message.text
    if text.lower() == 'отмена':
        await message.answer('Отменено', reply_markup=ReplyKeyboardRemove())
        await state.finish()  # обязательно завершаем состояние
    else:
        number_bid = message.text
        await state.update_data(number_bid=number_bid)

        # Создаем клавиатуру с кнопками статуса заявки
        keyboard_bid = types.ReplyKeyboardMarkup(resize_keyboard=True)
        roles = ['одобрена', 'отклонена', 'в обработке', 'отмена']
        keyboard_bid.add(*roles)

        await message.answer(f'Установите статус заявки #{number_bid}:', reply_markup=keyboard_bid)
        await SendMessage.status.set()  # устанавливаем состояние


@dp.message_handler(state=SendMessage.status)
async def input_status(message: types.Message, state: FSMContext):
    text = message.text
    if text.lower() == 'отмена':
        await message.answer('Отменено', reply_markup=ReplyKeyboardRemove())
        await state.finish()  # обязательно завершаем состояние
    else:
        status = message.text
        data = await state.get_data()
        number = int(data['number_bid'])

        await state.finish()
        # убрать кнопки
        keyboard = types.ReplyKeyboardRemove()

        # функция записи данных в БД
        if status in ["одобрена", "отклонена", "в обработке"]:
            await update_status_by_id(number, status)
            await message.answer(f'Для заявки #{number} установлено значение: "{status}"')
        else:
            await message.answer(f'Заявки с №{number} нет!', reply_markup=keyboard)


# /edit_bid_options - редактировать варианты заявок
@dp.message_handler(AdminsMessage(), Command('edit_bid_options'))
async def edit_application_options(message: types.Message):
    options = await get_variants_proposal()
    await message.answer(f'Варианты заявок: {options}, введите новые значения через запятую:',
                         reply_markup=keyboard_cancel)
    await SendMessage.options.set()  # устанавливаем состояние


# Отлавливаем состояние введенного сообщения
@dp.message_handler(state=SendMessage.options)
async def input_options(message: types.Message, state: FSMContext):
    text = message.text
    if text.lower() == 'отмена':
        await message.answer('Отменено', reply_markup=ReplyKeyboardRemove())
        await state.finish()  # обязательно завершаем состояние
    else:
        options = message.text
        output_list = options.split(', ')
        await set_variants_proposal(output_list)
        await message.answer('Установлено!')
        await state.finish()


# /list_users - выгрузить список пользователей
@dp.message_handler(AdminsMessage(), Command('list_users'))
async def list_users(message: types.Message):
    users = await select_all_users_with_data()
    # сохраняем в excel
    await save_to_excel(users)


from openpyxl.utils import get_column_letter


async def save_to_excel(data, file_name="пользователи.xlsx"):
    # Создайте новую книгу и выберите активный лист.
    wb = Workbook()
    sheet = wb.active

    # Напишите заголовки столбцов
    headers = ['user_id', 'Кто пригласил', 'first_name', 'last_name', 'username', 'Фамилия Имя Отчество', 'Город',
               'Номер телефона', 'Реферальная ссылка', 'Список пользователей зарегистрированных по реферальной ссылке',
               'Бонус 1', 'Бонус 2', 'Сумма на счету для вывода', 'Статус (роль)', 'Баланс (подписка)', 'status',
               'Количество рефералов по уровням', 'Дата регистрации']

    sheet.append(headers)

    # Установите ширину столбца 25 для всех столбцов.
    for col_num in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_num)
        sheet.column_dimensions[col_letter].width = 25

    # Запись данных на лист
    for row_data in data:
        sheet.append([row_data[header] for header in headers])

    # Сохраните книгу в файл
    wb.save(file_name)
    await send_list_users(dp, file_name)


async def send_list_users(dpr: Dispatcher, file):
    for admin in admins:
        try:
            await dpr.bot.send_message(chat_id=admin, text=f'Список пользователей:"')
            with open(file, 'rb') as document:
                await dpr.bot.send_document(chat_id=admin, document=InputFile(document))
        except Exception as err:
            logger.exception(err)


# /unload_bid - выгрузить заявки
@dp.message_handler(AdminsMessage(), Command('unload_bid'))
async def input_applications(message: types.Message):
    # Создаем клавиатуру с кнопками статуса заявок
    keyboard_applications = types.ReplyKeyboardMarkup(resize_keyboard=True)
    roles = ['одобрена', 'отклонена', 'в обработке', 'все заявки', 'отмена']
    keyboard_applications.add(*roles)
    await message.answer('Выберите какие заявки выгрузить:', reply_markup=keyboard_applications)
    await SendMessage.applications_selection.set()  # устанавливаем состояние


# Отлавливаем состояние applications_selection
@dp.message_handler(state=SendMessage.applications_selection)
async def input_applications_selection(message: types.Message, state: FSMContext):
    text = message.text
    if text.lower() == 'отмена':
        await message.answer('Отменено', reply_markup=ReplyKeyboardRemove())
        await state.finish()  # обязательно завершаем состояние
    else:
        applications_selection = message.text
        await state.update_data(applications_selection=applications_selection)
        valid_selections = ["одобрена", "отклонена", "в обработке", "все заявки"]
        get_applications = []
        if applications_selection in valid_selections:
            if applications_selection == "все заявки":
                get_applications = await get_all_proposals()
            else:
                get_applications = await get_proposals_by_status(applications_selection)
        else:
            await message.answer('Введено неверное значение')

        await message.answer(f'заявки со статусом "{applications_selection}":\n')
        await bid_out(get_applications)
        await state.finish()


async def bid_out(get_applications):
    application_dict = {}
    for application in get_applications:
        # Создаем словарь для текущей заявки
        current_application = {
            'user_id': application.user_id,
            'fio': application.fio,
            'variant_proposal': application.variant_proposal,
            'status_proposal': application.status_proposal,
            'approved_amount': application.approved_amount,
            'loan_amount': application.loan_amount
        }

        # Добавляем в application_dict с ключом in_id
        application_dict[application.id] = current_application

        # Возвращаем получившийся словарь
    await save_bid_to_excel(application_dict)


async def save_bid_to_excel(data, file_name='Заявки пользователей.xlsx'):
    # Создайте новую книгу и выберите активный лист.
    wb = Workbook()
    ws = wb.active

    aliases = ['номер заявки', 'user_id', 'Фамилия Имя Отчество', 'вариант заявки', 'статус заявки', 'одобренная сумма',
               'величина взятого займа']
    # Напишите заголовки столбцов
    ws.append(aliases)

    # Установите ширину столбца 25 для всех столбцов.
    for col_num in range(1, len(aliases) + 1):
        col_letter = get_column_letter(col_num)
        ws.column_dimensions[col_letter].width = 25

    # Запишите данные на рабочий лист
    for key, value in data.items():
        ws.append([key, value['user_id'], value['fio'], value['variant_proposal'], value['status_proposal'],
                   value['approved_amount'], value['loan_amount']])

    # Save the workbook
    wb.save(file_name)

    await send_bid(dp, file_name)


async def send_bid(dpr: Dispatcher, file):
    for admin in admins:
        try:
            with open(file, 'rb') as document:
                await dpr.bot.send_document(chat_id=admin, document=InputFile(document))
        except Exception as err:
            logger.exception(err)


# /replace_service_agreement - заменить договор оказания услуг
@dp.message_handler(AdminsMessage(), Command('replace_service_agreement'))
async def replace_service_agreement(message: types.Message):
    await message.answer(f'Загрузите новый договор оказания услуг:',
                         reply_markup=keyboard_cancel)
    await SendMessage.service_agreement.set()  # устанавливаем состояние


# Отлавливаем состояние введенного сообщения
@dp.message_handler(state=SendMessage.service_agreement,
                    content_types=[types.ContentType.DOCUMENT, types.ContentType.TEXT])
async def input_service_agreement(message: types.Message, state: FSMContext):
    if message.content_type == 'document':
        file_name = f'Договор оказания услуг.pdf'
        # Сохраняем
        file_path = os.path.join("admin_documents/договор оказания услуг", file_name)
        await message.document.download(destination_file=file_path)
        await message.answer('Новый договор успешно сохранен!')
        await state.finish()
    elif message.text.lower() == 'отмена':
        await message.answer('Отменено', reply_markup=ReplyKeyboardRemove())
        await state.finish()  # обязательно завершаем состояние

    else:
        await message.answer('Пожалуйста, загрузите документ:')
