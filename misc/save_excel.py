import os
import requests
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
from PIL import Image as PILImage
from loader import dp, bot
from bot_send.send_proposal import send_proposal
from data.config import BOT_TOKEN
from datetime import datetime
from loguru import logger
import io
import fitz  # PyMuPDF


async def download_telegram_image(file_id, output_path):
    bot_token = BOT_TOKEN  # Замените на свой токен
    api_url = f'https://api.telegram.org/bot{bot_token}/getFile'
    params = {'file_id': file_id}
    response = requests.get(api_url, params=params)
    file_path = response.json()['result']['file_path']
    file_url = f'https://api.telegram.org/file/bot{bot_token}/{file_path}'
    image_data = requests.get(file_url).content
    with open(output_path, 'wb') as file:
        file.write(image_data)


async def get_image_size(image_path):
    with PILImage.open(image_path) as img:
        return img.size


async def get_today_date():
    today_date = datetime.now().strftime('%d.%m.%Y')
    return today_date


async def create_excel_file(user_id, data):
        id_proposal = data['id']
        city = data['city']
        username = data['username']
        fio = data['fio']
        short_name = format_name(fio)
        variant_proposal = data['variant_proposal']
        today_date = await get_today_date()

        file_name = f'Заявка #{id_proposal}_{short_name}.xlsx'
        wb = Workbook()
        ws_user_data = wb.active
        ws_user_data.title = 'Данные пользователя'

        headers = ['user_id', 'username', 'ФИО', 'город', 'телефон', 'вид заявки']
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            ws_user_data[f'{col_letter}1'] = header

        # установка ширины столбцов
        for col_num in range(1, len(headers) + 1):
            col_letter = get_column_letter(col_num)
            ws_user_data.column_dimensions[col_letter].width = 30

        user_data = [user_id, username, data['fio'], data['city'], data['phone'], data['variant_proposal']]
        for col_num, value in enumerate(user_data, 1):
            col_letter = get_column_letter(col_num)
            ws_user_data[f'{col_letter}2'] = value

        photo_passport_1_path = f'temp/photo_passport_1_{fio}.jpg'
        await download_telegram_image(data['photo_passport_1'], photo_passport_1_path)
        img_passport_1 = Image(photo_passport_1_path)

        # Создайте новый лист для photo_passport_1.
        ws_passport_1 = wb.create_sheet(title='фото паспорта 1 страница')
        ws_passport_1.add_image(img_passport_1, 'A1')

        # Добавляем путь файла в список временных файлов
        temp_files = [photo_passport_1_path]

        photo_passport_2_path = f'temp/photo_passport_2_{fio}.jpg'
        await download_telegram_image(data['photo_passport_2'], photo_passport_2_path)
        img_passport_2 = Image(photo_passport_2_path)

        # Создайте новый лист для photo_passport_2.
        ws_passport_2 = wb.create_sheet(title='фото паспорта 2 страница')
        ws_passport_2.add_image(img_passport_2, 'A1')

        # Добавляем путь файла в список временных файлов
        temp_files.append(photo_passport_2_path)

        photo_snils_path = f'temp/photo_snils_{fio}.jpg'
        await download_telegram_image(data['photo_snils'], photo_snils_path)
        img_snils = Image(photo_snils_path)

        # Создайте новый лист для СНИЛС.
        ws_snils = wb.create_sheet(title='СНИЛ или вод.права')
        ws_snils.add_image(img_snils, 'A1')

        # Добавляем путь файла в список временных файлов
        temp_files.append(photo_snils_path)

        ws_credit_history_1 = wb.create_sheet(title='Данные с 1 сайта кредитных историй')
        photo_ids_1 = data['photo_from_1_credit_history_site']
        document = photo_ids_1[0]
        if document.startswith("AgAC"):  # если картинка
            img_size_indent = 1
            for row_num, photo_id in enumerate(photo_ids_1, 1):
                photo_path = f'temp/photo_credit_history_1_{fio}_{row_num}.jpg'
                await download_telegram_image(photo_id, photo_path)
                img_credit_history_1 = Image(photo_path)
                ws_credit_history_1.add_image(img_credit_history_1, f'A{img_size_indent}')
                get_size = await get_image_size(photo_path)
                img_size_indent += round((get_size[1]) / 20)

                # Добавляем путь файла в список временных файлов
                temp_files.append(photo_path)
        elif document.startswith("temp"):  # если pdf файл
            try:
                path = f"temp/{user_id}_document_1.pdf"
                pdf_document = fitz.open(path)

                # Проходим по страницам PDF и преобразуем их в изображения
                img_size_indent = 1
                for page_number in range(pdf_document.page_count):
                    page = pdf_document[page_number]
                    image = page.get_pixmap()
                    image_pil = PILImage.frombytes("RGB", [image.width, image.height], image.samples)
                    photo_path = f"{fio}_pdf_1_page_{page_number + 1}.png"
                    image_pil.save(photo_path)
                    img_credit_history_1 = Image(photo_path)
                    ws_credit_history_1.add_image(img_credit_history_1, f'A{img_size_indent}')
                    get_size = await get_image_size(photo_path)
                    img_size_indent += round((get_size[1]) / 20)

                    # Добавляем путь файла в список временных файлов
                    temp_files.append(photo_path)
                pdf_document.close()
                temp_files.append(path)

            except Exception as e:
                logger.exception(e)

        ws_credit_history_2 = wb.create_sheet(title='Данные со 2 сайта кредитных историй')
        photo_ids_2 = data['photo_from_2_credit_history_site']
        photo_2 = photo_ids_2[0]
        if photo_2.startswith("AgAC"):  # если картинка
            img_size_indent = 1
            for row_num, photo_id in enumerate(photo_ids_2, 1):
                photo_path = f'temp/photo_credit_history_2_{fio}_{row_num}.jpg'
                await download_telegram_image(photo_id, photo_path)
                img_credit_history_2 = Image(photo_path)
                ws_credit_history_2.add_image(img_credit_history_2, f'A{img_size_indent}')
                get_size = await get_image_size(photo_path)
                img_size_indent += round((get_size[1]) / 20)

                # Добавляем путь файла в список временных файлов
                temp_files.append(photo_path)
        elif document.startswith("temp"):  # если pdf файл
            try:
                path = f"temp/{user_id}_document_2.pdf"
                pdf_document = fitz.open(path)

                # Проходим по страницам PDF и преобразуем их в изображения
                img_size_indent = 1
                for page_number in range(pdf_document.page_count):
                    page = pdf_document[page_number]
                    image = page.get_pixmap()
                    image_pil = PILImage.frombytes("RGB", [image.width, image.height], image.samples)
                    photo_path = f"{fio}_pdf_2_page_{page_number + 1}.png"
                    image_pil.save(photo_path)
                    img_credit_history_2 = Image(photo_path)
                    ws_credit_history_2.add_image(img_credit_history_2, f'A{img_size_indent}')
                    get_size = await get_image_size(photo_path)
                    img_size_indent += round((get_size[1]) / 20)

                    # Добавляем путь файла в список временных файлов
                    temp_files.append(photo_path)
                pdf_document.close()
                temp_files.append(path)

            except Exception as e:
                logger.exception(e)

        # Сохраняем файл
        path = f'admin_documents/proposal/{file_name}'
        wb.save(path)

        # Удаляем временные файлы
        for temp_file in temp_files:
            os.remove(temp_file)

        await send_proposal(dp, username, path, today_date, fio, city, variant_proposal)


def format_name(input_string):
    parts = input_string.split()

    if len(parts) == 3:
        return f"{parts[0]} {parts[1][0]}.{parts[2][0]}."
    else:
        return input_string